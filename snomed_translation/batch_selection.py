"""Reproducible, feedback-aware selection of a new SME review batch."""
from __future__ import annotations
import csv, hashlib, json, math, re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from pipelines.context import RunContext
from pipelines.functions import FunctionResult

import logging
log = logging.getLogger(__name__)

STOP = {"of", "the", "and", "with", "without", "using", "for", "to", "in",
        "on", "by", "a", "an", "procedure", "imaging"}
LATERALITY = {"left", "right", "bilateral", "upper", "lower"}
TOPIC_STOP = {
    "computed", "tomography", "ct", "magnetic", "resonance", "mri",
    "ultrasound", "ultrasonography", "ultrasonographic", "echography",
    "radiography", "radiographic", "x", "ray", "fluoroscopy", "fluoroscopic",
    "angiography", "angiographic", "venography", "radionuclide", "positron",
    "emission", "pet", "spect", "contrast", "guidance", "guided", "study",
}

def path_of(value: Any) -> str | None:
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        for key in ("_primary", "dataset", "rows", "path"):
            if isinstance(value.get(key), str):
                return value[key]
    return None

def read_rows(value: Any) -> list[dict]:
    path = path_of(value)
    if not path or not Path(path).exists():
        return []
    with Path(path).open(encoding="utf-8") as handle:
        return list(csv.DictReader(handle))

def tokens(text: str, drop_laterality: bool = False) -> set[str]:
    values = set(re.findall(r"[a-z0-9]+", text.lower())) - STOP
    return values - LATERALITY if drop_laterality else values

def similarity(a: str, b: str) -> float:
    left, right = tokens(a, True), tokens(b, True)
    return len(left & right) / max(1, len(left | right))

def topic_tokens(text: str) -> set[str]:
    """Non-modality content used to prevent one anatomy/template dominating."""
    return tokens(text, True) - TOPIC_STOP

def modality(text: str) -> str:
    value = text.lower()
    rules = [("ct", r"computed tomography|\bct\b"),
             ("mri", r"magnetic resonance|\bmri\b"),
             ("ultrasound", r"ultrasound|ultrason|echograph"),
             ("nuclear", r"scintigraph|radionuclide|positron emission|\bpet\b|spect"),
             ("radiography", r"radiograph|x-ray|fluoroscop"),
             ("angiography", r"angiograph|arteriograph|venograph")]
    return next((name for name, pattern in rules if re.search(pattern, value)), "other")

def complexity(text: str) -> str:
    n = len(text.split())
    return "short" if n <= 4 else ("medium" if n <= 8 else "long")

def stable(row: dict, seed: int) -> str:
    return hashlib.sha256(f"{seed}:{row.get('sctid', '')}".encode()).hexdigest()

def select_sme_batch(ctx: RunContext, inputs: dict[str, Any],
                     params: dict[str, Any]) -> FunctionResult:
    candidates = read_rows(inputs.get("candidates"))
    previous = read_rows(inputs.get("previous_sme"))
    backtrans = read_rows(inputs.get("backtranslations"))
    if not candidates or not previous:
        return FunctionResult(ok=False, message="select_sme_batch needs candidates and previous_sme")
    strategy = str(params.get("strategy") or "balanced").lower()
    size, seed = int(params.get("size") or 100), int(params.get("seed") or 20260712)
    max_prior = float(params.get("max_prior_similarity") or 0.72)
    max_internal = float(params.get("max_internal_similarity") or 0.50)
    max_topic = max(0, int(params.get("max_topic_repeats") or 10))
    previous_ids = {(row.get("sctid") or "").strip() for row in previous}
    previous_terms = [(row.get("english_term") or row.get("preferred_term") or "").strip()
                      for row in previous]
    bt = {}
    for row in backtrans:
        try:
            bt[(row.get("sctid") or "").strip()] = float(row.get("sim_en_back") or 1)
        except ValueError:
            pass
    bad, good = Counter(), Counter()
    for row in previous:
        target = bad if (row.get("sme_rating") or "").upper() != "ACCEPTABLE" else good
        # Laterality/location words describe where a prior correction happened,
        # not necessarily why it was wrong. Including them made one earlier batch
        # select 27 upper-limb terms after the SME corrected 위팔 -> 팔/상지.
        target.update(tokens(row.get("english_term") or "", True))
    feedback = {word: math.log((bad[word] + 1) / (good[word] + 1))
                for word in bad if bad[word] >= 2}
    eligible, near_prior = [], 0
    for row in candidates:
        sctid = (row.get("sctid") or "").strip()
        english = (row.get("preferred_term") or row.get("english") or "").strip()
        korean = (row.get("translation") or "").strip()
        if not sctid or not english or not korean or sctid in previous_ids:
            continue
        nearest = max((similarity(english, term) for term in previous_terms), default=0)
        if nearest > max_prior:
            near_prior += 1
            continue
        sim = bt.get(sctid, 1.0)
        risk = 2 * (1 - sim) + min(len(english.split()), 14) / 20
        risk += sum(max(0, feedback.get(word, 0)) for word in tokens(english)) / 4
        risk += 0.25 if modality(english) == "other" else 0
        eligible.append({**row, "preferred_term": english, "translation": korean,
                         "modality_group": modality(english),
                         "complexity_bin": complexity(english),
                         "sim_en_back": sim, "selection_risk": round(risk, 4),
                         "nearest_batch1_jaccard": round(nearest, 4)})
    def add(pool: list[dict], n: int, arm: str, chosen: list[dict],
            *, enforce_topic_cap: bool = True,
            similarity_limit: float | None = None) -> None:
        target = len(chosen) + n
        limit = max_internal if similarity_limit is None else similarity_limit
        topic_counts = Counter(token for item in chosen
                               for token in topic_tokens(item["preferred_term"]))
        for row in pool:
            if len(chosen) >= target:
                break
            if any(similarity(row["preferred_term"], item["preferred_term"]) >= limit
                   for item in chosen):
                continue
            row_topics = topic_tokens(row["preferred_term"])
            if (enforce_topic_cap and max_topic and
                    any(topic_counts[token] >= max_topic for token in row_topics)):
                continue
            chosen.append({**row, "selection_arm": arm})
            topic_counts.update(row_topics)
    cells = defaultdict(list)
    for row in eligible:
        cells[(row["modality_group"], row["complexity_bin"])].append(row)
    for rows in cells.values():
        rows.sort(key=lambda row: stable(row, seed))
    coverage, keys = [], sorted(cells)
    while any(cells.values()):
        for key in keys:
            if cells[key]:
                coverage.append(cells[key].pop(0))
    # Risk-sort within modality/complexity cells, then round-robin across cells.
    # This retains the active signal without letting one error-associated term
    # family consume the entire active arm.
    active_cells = defaultdict(list)
    for row in eligible:
        active_cells[(row["modality_group"], row["complexity_bin"])].append(row)
    for rows in active_cells.values():
        rows.sort(key=lambda row: (-row["selection_risk"], stable(row, seed)))
    active, active_keys = [], sorted(active_cells)
    while any(active_cells.values()):
        for key in active_keys:
            if active_cells[key]:
                active.append(active_cells[key].pop(0))
    chosen: list[dict] = []
    if strategy == "coverage":
        add(coverage, size, "coverage", chosen)
    elif strategy == "active":
        add(active, size, "active", chosen)
    else:
        add(active, size // 2, "active", chosen)
        add(coverage, size - len(chosen), "coverage", chosen)
    if len(chosen) < size:
        add(sorted(eligible, key=lambda row: stable(row, seed + 1)),
            size - len(chosen), "fallback", chosen)
    if len(chosen) < size:
        # A deterministic escape hatch for unusually small pools. Keep the
        # stricter within-batch similarity threshold, but relax topic caps.
        add(sorted(eligible, key=lambda row: stable(row, seed + 2)),
            size - len(chosen), "fallback_relaxed", chosen,
            enforce_topic_cap=False)
    chosen = chosen[:size]
    if len(chosen) < size:
        return FunctionResult(ok=False, message=f"only {len(chosen)} sufficiently novel candidates")
    output = Path(ctx.log_dir) / f"sme_batch_{strategy}_{size}.csv"
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = ["sctid", "preferred_term", "translation", "selection_arm",
              "modality_group", "complexity_bin", "sim_en_back", "selection_risk",
              "nearest_batch1_jaccard", "sme_rating", "sme_corrected_ko",
              "sme_error_category", "sme_notes"]
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(chosen)
    metrics = {"pool_rows": float(len(candidates)), "eligible_novel": float(len(eligible)),
               "excluded_previous": float(len(previous_ids)),
               "excluded_near_prior": float(near_prior), "selected": float(len(chosen)),
               "mean_backtrans_similarity": round(sum(row["sim_en_back"] for row in chosen) / size, 4),
               "mean_selection_risk": round(sum(row["selection_risk"] for row in chosen) / size, 4)}
    pair_similarities = [similarity(left["preferred_term"], right["preferred_term"])
                         for index, left in enumerate(chosen)
                         for right in chosen[index + 1:]]
    selected_topics = Counter(token for row in chosen
                              for token in topic_tokens(row["preferred_term"]))
    metrics["selected_max_pair_similarity"] = round(max(pair_similarities, default=0), 4)
    metrics["selected_max_topic_repeats"] = float(max(selected_topics.values(), default=0))
    for label, count in Counter(row["selection_arm"] for row in chosen).items():
        metrics[f"arm_{label}"] = float(count)
    for label, count in Counter(row["modality_group"] for row in chosen).items():
        metrics[f"modality_{label}"] = float(count)
    for label, count in Counter(row["complexity_bin"] for row in chosen).items():
        metrics[f"complexity_{label}"] = float(count)
    audit = Path(ctx.log_dir) / f"sme_batch_{strategy}_audit.json"
    audit.write_text(json.dumps({"strategy": strategy, "seed": seed,
                                 "selection_parameters": {
                                     "max_prior_similarity": max_prior,
                                     "max_internal_similarity": max_internal,
                                     "max_topic_repeats": max_topic,
                                 },
                                 "feedback_terms": feedback,
                                 "selected_topic_counts": dict(selected_topics.most_common()),
                                 "metrics": metrics},
                                ensure_ascii=False, indent=2), encoding="utf-8")
    return FunctionResult(ok=True, outputs={"batch": str(output), "audit": str(audit)},
                          metrics=metrics, message=f"selected {len(chosen)} novel imaging procedures")


def package_sme_batch(ctx: RunContext, inputs: dict[str, Any],
                      params: dict[str, Any]) -> FunctionResult:
    """Join selection metadata to newly generated translations for SME review."""
    selected = read_rows(inputs.get("selection"))
    translated = read_rows(inputs.get("translations"))
    if not selected or not translated:
        return FunctionResult(
            ok=False,
            message="package_sme_batch needs selection and translations datasets",
        )
    id_col = str(params.get("id_col") or "sctid")
    translation_col = str(params.get("translation_col") or "translation")
    generated = {
        (row.get(id_col) or "").strip(): (row.get(translation_col) or "").strip()
        for row in translated
    }
    fields = [
        "sctid", "preferred_term", "translation", "selection_arm",
        "modality_group", "complexity_bin", "sim_en_back", "selection_risk",
        "nearest_batch1_jaccard", "sme_rating", "sme_corrected_ko",
        "sme_error_category", "sme_notes",
    ]
    rows = []
    missing = 0
    errors = 0
    for source in selected:
        sctid = (source.get(id_col) or "").strip()
        translation = generated.get(sctid, "")
        missing += int(not translation)
        errors += int(translation.startswith("ERROR:"))
        row = {field: source.get(field, "") for field in fields}
        row["sctid"] = sctid
        row["translation"] = translation
        for field in ("sme_rating", "sme_corrected_ko", "sme_error_category", "sme_notes"):
            row[field] = ""
        rows.append(row)

    output = Path(ctx.log_dir) / "sme_review_packet.csv"
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    metrics = {
        "packet_rows": float(len(rows)),
        "missing_translations": float(missing),
        "translation_errors": float(errors),
    }
    return FunctionResult(
        ok=missing == 0 and errors == 0,
        outputs={"packet": str(output)},
        metrics=metrics,
        message=(f"packaged {len(rows)} SME rows; missing={missing}, "
                 f"translation_errors={errors}"),
    )


def package_deliverable(ctx: RunContext, inputs: dict[str, Any],
                        params: dict[str, Any]) -> FunctionResult:
    """Assemble the reviewer-facing deliverable from a translation set.

    Distinct from ``package_sme_batch``, which packages a SAMPLED batch and is
    built around selection provenance (arm, modality group, risk) that a full
    deliverable has no analogue for. Shared where it matters: the reviewer's
    response columns keep the same names, so the spreadsheet looks identical to
    previous batches.

    Three joins, in this order:

    1. SME OVERLAY. Where a concept has an adjudicated rendering it wins over
       the machine's, and its accepted synonyms come with it. This is not a
       nicety — the deliverable is assembled from raw cascade output, so
       without it 130 reviewed rows ship as machine text and two of them ship
       as blockers whose adjudicated form is clean.
    2. PRIORITY. review_priority comes from the qa_gate worklist so the
       reviewer's attention follows the detectors, and `done` marks the
       adjudicated rows so nobody re-reviews them.
    3. PRUNE. Columns listed in ``drop_cols`` are omitted. Signals that fired
       on nothing (an all-empty judge label) or fired once as a known false
       positive are noise in a reviewer's spreadsheet, and a column of blanks
       invites the reader to wonder what it was supposed to mean.
    """
    translations = read_rows(inputs.get("translations"))
    if not translations:
        return FunctionResult(ok=False, message="package_deliverable needs `translations`")
    gold = {(r.get("sctid") or "").strip(): r for r in read_rows(inputs.get("gold"))}
    priority = {(r.get("sctid") or "").strip(): r
                for r in read_rows(inputs.get("priority"))}

    ko_col = str(params.get("ko_col") or "translation")
    gold_col = str(params.get("gold_col") or "ko_reference")

    # An adjudicated answer is authoritative but NOT immortal: a later ruling
    # supersedes an earlier row. All 7 gold rows rendering plain radiography as
    # 일반 x선 are batch1; all 4 using the ruled form 단순 촬영 are batch2. So
    # overlaying gold blindly would have re-introduced the very forms a later
    # SME ruling replaced — and several of those rows are rated only PARTIAL.
    # The overlay therefore obeys the same law as every other patch here: it
    # may not introduce a blocker. Where it would, the machine text stands and
    # the row goes to the reviewer as a conflict to settle.
    from snomed_translation.hard_rules import find_violations, load_hard_rules
    blocker_rules = [r for r in load_hard_rules(params.get("rules_file"))
                     if r.severity == "blocker"] if params.get("rules_file") else []
    drop = {c.strip() for c in str(params.get("drop_cols") or "").split(",") if c.strip()}
    review_cols = ["sme_rating", "sme_corrected_ko", "sme_error_category", "sme_notes"]

    fields = ["sctid", "preferred_term", "translation_ko", "synonyms_ko",
              "translation_source", "review_priority", "flagged_checks",
              "n_distinct", "routed"] + review_cols
    fields = [f for f in fields if f not in drop]

    rows, n_gold, n_high, n_superseded = [], 0, 0, 0
    for r in translations:
        sctid = (r.get("sctid") or "").strip()
        en = (r.get("preferred_term") or "").strip()
        g = gold.get(sctid)
        if g and blocker_rules and find_violations(
                (g.get(gold_col) or "").strip(), blocker_rules,
                require_enforce=False, source=en):
            g = None
            n_superseded += 1
        if g:
            n_gold += 1
            translation = (g.get(gold_col) or "").strip()
            syns = [s for s in (g.get("ko_all") or "").split("|")[1:] if s.strip()]
            source, prio = "sme_approved", "done"
            if (g.get("sme_rating") or "").strip().upper() == "PARTIAL":
                # Rated partially correct, so not a settled answer: ship it but
                # send it back rather than marking it done.
                source, prio = "sme_partial", "high"
                n_high += 1
        else:
            translation = (r.get(ko_col) or "").strip()
            syns, source = [], "machine_v6_0"
            p = priority.get(sctid)
            # Disagreement is the strongest single predictor we have of an
            # incorrect row (unanimous 57.7% exact vs 14.6% when the samples
            # disagree, AUC 0.755), so it drives priority alongside the
            # detectors rather than being relegated to a column.
            nd = int(r.get("n_distinct") or 0)
            if p and (p.get("max_severity") or "") == "blocker":
                prio = "high"
            elif p and int(p.get("n_checks") or 1) > 1:
                prio = "high"
            elif nd >= 4:
                prio = "high"
            elif p or nd >= 2:
                prio = "medium"
            else:
                prio = "low"
            n_high += int(prio == "high")
        row = {
            "sctid": sctid,
            "preferred_term": (r.get("preferred_term") or "").strip(),
            "translation_ko": translation,
            "synonyms_ko": " | ".join(syns),
            "translation_source": source,
            "review_priority": prio,
            "flagged_checks": (priority.get(sctid, {}).get("checks") or ""),
            "n_distinct": r.get("n_distinct", ""),
            "routed": r.get("routed", ""),
        }
        row.update({c: "" for c in review_cols})
        rows.append({k: v for k, v in row.items() if k in fields})

    out = Path(ctx.artifacts_dir() or ctx.log_dir) / (
        f"{params.get('output_name') or 'deliverable'}.csv")
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    # Pre-drawn stratified sample. The point of review_priority is to predict
    # where the errors are, and that claim is only testable if the reviewer
    # looks at all three tiers — reviewing only `high` measures nothing. Drawing
    # the sample here rather than asking the reviewer to filter also removes
    # two ways the comparison could go wrong: contamination of `high` by the 93
    # sme_partial rows (which are not machine output and would bias its error
    # rate), and selection within a tier toward rows that look interesting.
    # Deterministic in sctid, so the same pack always yields the same sample.
    per_tier = int(params.get("sample_per_tier") or 0)
    n_sampled = 0
    if per_tier:
        by_tier = defaultdict(list)
        for row in rows:
            if row.get("translation_source") == "machine_v6_0":
                by_tier[row.get("review_priority")].append(row)
        picked = set()
        for tier in ("high", "medium", "low"):
            pool = sorted(by_tier.get(tier, []),
                          key=lambda r: stable(r, int(params.get("seed") or 20260812)))
            picked.update(id(r) for r in pool[:per_tier])
        for row in rows:
            row["review_sample"] = "yes" if id(row) in picked else ""
        n_sampled = sum(1 for r in rows if r.get("review_sample") == "yes")
        if "review_sample" not in fields:
            fields.insert(fields.index("review_priority") + 1, "review_sample")
        with out.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

    missing = sum(1 for r in rows if not r.get("translation_ko"))

    xlsx_info = {}
    if params.get("xlsx"):
        xlsx_path = out.with_suffix(".xlsx")
        try:
            xlsx_info = _write_review_xlsx(out, xlsx_path, fields)
        except Exception as exc:  # noqa: BLE001 — the CSV is the artifact of record
            log.warning("package_deliverable: xlsx rendering failed (%s); "
                        "the CSV is unaffected", exc)

    return FunctionResult(
        ok=missing == 0,
        outputs={"deliverable": str(out), "dataset": str(out),
                 **({"xlsx": str(out.with_suffix(".xlsx"))} if xlsx_info else {})},
        metrics={"n_rows": float(len(rows)), "n_sme_approved": float(n_gold),
                 "n_high_priority": float(n_high),
                 "n_gold_superseded": float(n_superseded),
                 "n_missing_translation": float(missing),
                 "n_columns": float(len(fields)),
                 "n_review_sample": float(n_sampled),
                 "xlsx_written": 1.0 if xlsx_info else 0.0},
        message=(f"packaged {len(rows)} rows: {n_gold} SME-approved (marked "
                 f"done), {n_superseded} adjudicated rows withheld as "
                 f"superseded by a later ruling, {n_high} high priority, "
                 f"{missing} missing; {n_sampled} drawn for the priority "
                 f"experiment; {len(fields)} columns"))


# Detector ids are internal vocabulary. A clinician filtering `flagged_checks`
# should not meet the word "hallucinated" — it reads as an accusation about a
# term they may have written, and it invites them to audit our detectors rather
# than the translations. Anything unmapped degrades to a neutral phrase.
CHECK_LABELS = {
    "hierarchy-inconsistency": "differs from parent concept",
    "hallucinated": "contrast wording may be added",
    "dropped": "contrast wording may be missing",
    "transliteration": "may be transliterated rather than translated",
    "no-rf2-markup": "source formatting may have carried over",
    "upper-limb-not-upper-arm": "body-site term to confirm",
    "repeated-token": "wording repeats",
}


def _label_checks(value: str) -> str:
    out = []
    for tok in (value or "").split(";"):
        tok = tok.strip()
        if tok:
            out.append(CHECK_LABELS.get(tok, "flagged for checking"))
    return "; ".join(dict.fromkeys(out))


def _write_review_xlsx(csv_path: Path, xlsx_path: Path, fields: list[str]) -> dict:
    """Render the review pack as a workbook a reviewer can actually work in.

    The CSV is the machine-readable artifact; this is the human one. Reviewers
    open it, filter, and type into it for hours, so the affordances are the
    deliverable: a frozen header and ID/term columns so context never scrolls
    away, an autofilter so they can slice by priority themselves, priority
    colour-coded so a glance locates the work, and the four response columns
    visually separated so it is obvious where to type.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with csv_path.open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.append(fields)

    header_fill = PatternFill("solid", fgColor="1F3864")
    response_fill = PatternFill("solid", fgColor="7030A0")
    response_cols = {"sme_rating", "sme_corrected_ko", "sme_error_category", "sme_notes"}
    for i, name in enumerate(fields, 1):
        c = ws.cell(row=1, column=i)
        c.fill = response_fill if name in response_cols else header_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=True)

    prio_fill = {
        "high": PatternFill("solid", fgColor="F8CBAD"),
        "medium": PatternFill("solid", fgColor="FFE699"),
        "low": PatternFill("solid", fgColor="E2EFDA"),
        "done": PatternFill("solid", fgColor="D9D9D9"),
    }
    prio_idx = fields.index("review_priority") + 1 if "review_priority" in fields else None
    for r in rows:
        ws.append([_label_checks(r.get(k, "")) if k == "flagged_checks"
                   else r.get(k, "") for k in fields])
        if prio_idx:
            cell = ws.cell(row=ws.max_row, column=prio_idx)
            fill = prio_fill.get((r.get("review_priority") or "").strip())
            if fill:
                cell.fill = fill

    widths = {"sctid": 16, "preferred_term": 52, "translation_ko": 40,
              "synonyms_ko": 26, "translation_source": 17, "review_priority": 15,
              "flagged_checks": 26, "n_distinct": 11, "routed": 12,
              "sme_rating": 14, "sme_corrected_ko": 34,
              "sme_error_category": 20, "sme_notes": 40}
    for i, name in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 18)
    # D2, not C2: the reviewer types in the last four columns, and with C2 the
    # Korean they are correcting scrolls off-screen exactly when they need it.
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    # Constrain the rating so the returned file needs no repair. Free text has
    # previously come back as CORRECT, blanks, and mixed case, which the gold
    # builder then has to reconcile heuristically.
    from openpyxl.worksheet.datavalidation import DataValidation
    if "sme_rating" in fields:
        dv = DataValidation(type="list",
                            formula1='"ACCEPTABLE,PARTIAL,WRONG"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(fields.index("sme_rating") + 1)
        dv.add(f"{col}2:{col}{ws.max_row}")

    # BLINDED SAMPLE SHEET. The 120 rows exist to test whether review_priority
    # predicts errors — and on the Review sheet each one displays its own
    # priority, n_distinct and flags while being rated. A row labelled `high`
    # invites scrutiny and `low` invites a wave-through, which would manufacture
    # exactly the result we are trying to test. This sheet shows the same rows
    # with those cues removed, in an order that does not group by tier, so the
    # rating is formed from the translation alone. Merged back by sctid.
    sample = [r for r in rows if (r.get("review_sample") or "").strip() == "yes"]
    if sample:
        blind_fields = ["sctid", "preferred_term", "translation_ko", "synonyms_ko",
                        "sme_rating", "sme_corrected_ko", "sme_error_category",
                        "sme_notes"]
        blind_fields = [f for f in blind_fields if f in fields]
        # Deterministic shuffle: order by a hash of the sctid so tiers interleave
        # reproducibly rather than appearing in blocks.
        sample = sorted(sample, key=lambda r: hashlib.sha256(
            f"blind:{r.get('sctid','')}".encode()).hexdigest())
        sh = wb.create_sheet("Sample - please do these first", 1)
        sh.append(blind_fields)
        for i, name in enumerate(blind_fields, 1):
            c = sh.cell(row=1, column=i)
            c.fill = response_fill if name in response_cols else header_fill
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r in sample:
            sh.append([r.get(k, "") for k in blind_fields])
        for i, name in enumerate(blind_fields, 1):
            sh.column_dimensions[get_column_letter(i)].width = widths.get(name, 18)
        sh.freeze_panes = "D2"
        sh.auto_filter.ref = sh.dimensions
        sh.row_dimensions[1].height = 30
        if "sme_rating" in blind_fields:
            dv2 = DataValidation(type="list",
                                 formula1='"ACCEPTABLE,PARTIAL,WRONG"', allow_blank=True)
            sh.add_data_validation(dv2)
            col = get_column_letter(blind_fields.index("sme_rating") + 1)
            dv2.add(f"{col}2:{col}{sh.max_row}")

    # A second sheet so the column semantics travel WITH the file. A legend in
    # a covering email is lost the moment the file is forwarded.
    key = wb.create_sheet("How to read this")
    for line in [
        ["Column", "Meaning"],
        ["translation_ko", "The proposed Korean term. This is what needs checking."],
        ["synonyms_ko", "Additional accepted forms already on record."],
        ["translation_source", "machine_v6_0 = newly generated. sme_approved = your previously accepted wording, unchanged. sme_partial = you rated it PARTIAL before, so it is back for another look."],
        ["review_priority", "Where our automated checks think the risk is: high, medium, low. done = previously accepted, no action needed."],
        ["flagged_checks", "Which check fired, when one did. hierarchy-inconsistency = rendered differently from its own parent concept."],
        ["n_distinct", "How many different answers the model gave across 5 attempts. 1 = it was consistent; 4-5 = it was unsure."],
        ["review_sample", "yes = one of the 120 rows we would most like done. They are laid out on the 'Sample - please do these first' sheet."],
        ["routed", "Which model produced the row. Internal bookkeeping; nothing for you to act on."],
        ["", ""],
        ["Please fill in", ""],
        ["sme_rating", "ACCEPTABLE / PARTIAL / WRONG"],
        ["sme_corrected_ko", "Your preferred wording, if it needs changing"],
        ["sme_error_category", "e.g. wrong referent, wrong register, word order, missing element"],
        ["sme_notes", "Anything else worth recording"],
    ]:
        key.append(line)
    key.column_dimensions["A"].width = 22
    key.column_dimensions["B"].width = 110
    for c in key["A"]:
        c.font = Font(bold=True)
    for row in key.iter_rows(min_col=2, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    key["A1"].fill = header_fill
    key["B1"].fill = header_fill
    key["A1"].font = key["B1"].font = Font(color="FFFFFF", bold=True)

    wb.save(xlsx_path)
    return {"n_rows": len(rows), "n_cols": len(fields)}
