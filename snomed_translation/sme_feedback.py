"""SME-feedback analysis nodes (batch-2, 2026-08).

Two function nodes:

- ``contrast_fidelity_detect`` — deterministic, source-conditional detector
  for the top SME "Wrong" class: a contrast phrase (조영제 사용 / 조영제
  미사용) in the Korean output that the English source does not license, or
  a source contrast modifier dropped from the output. Sibling of
  ``transliteration_detect`` (same flag-CSV + metrics shape).

- ``sme_metric_separation`` — scores each SME-reviewed translation against
  the SME gold (multi-reference) with the candidate metrics GEPA could
  optimise (spacing-normalised exact, chrF, BGE-M3 cosine) and measures how
  well each separates the SME's Correct/Acceptable rows from Partial/Wrong.
  The SME's stated boundary (spacing + preferred-synonym differences
  acceptable; word-order differences not) is what a good metric must
  reproduce.
"""
from __future__ import annotations

import csv
import logging
import os
import re
import time
from pathlib import Path
from threading import Lock
from typing import Any

import sacrebleu

from pipelines.context import RunContext
from pipelines.functions import FunctionResult

log = logging.getLogger(__name__)

from snomed_translation.evidence_analysis import (  # shared helpers
    _auc,
    _best_threshold,
    _embed_similarity,
)
from snomed_translation.scoring import norm_text

WITH_RE = re.compile(r"\bwith contrast\b", re.IGNORECASE)
WITHOUT_RE = re.compile(r"\bwithout contrast\b", re.IGNORECASE)
ANY_CONTRAST_RE = re.compile(r"contrast", re.IGNORECASE)

KO_WITH = "조영제 사용"
KO_WITHOUT = "조영제 미사용"

ACCEPT_RATINGS = {"CORRECT", "ACCEPTABLE"}
REJECT_RATINGS = {"PARTIAL", "WRONG"}


def _dataset_path(value: Any) -> str | None:
    if isinstance(value, dict):
        return value.get("path") or value.get("csv")
    if isinstance(value, str):
        return value
    return None


def _roles(value: Any) -> dict:
    if isinstance(value, dict):
        return value.get("roles") or {}
    return {}


def _col(params: dict, roles: dict, param_key: str, role_key: str,
         default: str) -> str:
    return str(params.get(param_key) or roles.get(role_key) or default)


# ---------------------------------------------------------------------------
# contrast_fidelity_detect
# ---------------------------------------------------------------------------


def contrast_issue(en: str, ko: str) -> str:
    """Deterministic contrast-fidelity verdict for one (source, output) pair.

    Returns "" (clean/ambiguous), "hallucinated", "wrong_polarity" or
    "dropped". Sources mentioning contrast outside the with/without modifier
    forms ("contrast procedure", "double contrast …") are treated as
    ambiguous and never flagged.
    """
    has_with = bool(WITH_RE.search(en))
    has_without = bool(WITHOUT_RE.search(en))
    if ANY_CONTRAST_RE.search(en) and not (has_with or has_without):
        return ""
    ko_without = KO_WITHOUT in ko
    ko_with = KO_WITH in ko
    if not (has_with or has_without):
        return "hallucinated" if "조영제" in ko else ""
    if has_with:
        if ko_without:
            return "wrong_polarity"
        return "" if ko_with else "dropped"
    if ko_with and not ko_without:
        return "wrong_polarity"
    return "" if ko_without else "dropped"


def contrast_fidelity_detect(ctx: RunContext, inputs: dict[str, Any],
                             params: dict[str, Any]) -> FunctionResult:
    """Flag contrast-phrase mismatches between source and output.

    Cases (deterministic):
      hallucinated  — source has NO occurrence of "contrast" at all, output
                      contains 조영제.
      wrong_polarity — source says "with contrast" but output has 조영제
                      미사용, or "without contrast" but output has 조영제
                      사용.
      dropped       — source says "with/without contrast" but the required
                      조영제 phrase is absent from the output.
    Sources that mention contrast in other constructions ("contrast
    procedure", "double contrast …") are skipped as ambiguous — the goal is
    high precision.
    """
    t0 = time.monotonic()
    tpath = _dataset_path(inputs.get("translations"))
    if not tpath or not Path(tpath).exists():
        return FunctionResult(
            ok=False, message="contrast_fidelity_detect: no `translations` dataset wired")
    roles = _roles(inputs.get("translations"))
    id_col = _col(params, roles, "id_col", "sctid", "sctid")
    en_col = _col(params, roles, "en_col", "en", "en")
    ko_col = _col(params, roles, "ko_col", "target", "translation")
    label_col = str(params.get("label_col") or "sme_rating")

    out_rows: list[dict] = []
    with Path(tpath).open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        has_label = label_col in (reader.fieldnames or [])
        for r in reader:
            en, ko = (r.get(en_col) or "").strip(), (r.get(ko_col) or "").strip()
            if not en or not ko:
                continue
            issue = contrast_issue(en, ko)
            row = {id_col: (r.get(id_col) or "").strip(), "english": en,
                   "korean": ko, "issue": issue, "flag": int(bool(issue))}
            if has_label:
                row["sme_rating"] = (r.get(label_col) or "").strip().upper()
            out_rows.append(row)

    if not out_rows:
        return FunctionResult(
            ok=False, message=f"contrast_fidelity_detect: no usable rows in {tpath} "
                              f"(en_col={en_col!r}, ko_col={ko_col!r})")

    out = Path(ctx.log_dir) / "contrast_fidelity_flags.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        w.writeheader()
        w.writerows(out_rows)

    n = len(out_rows)
    flagged = [r for r in out_rows if r["flag"]]
    by_issue: dict[str, int] = {}
    for r in flagged:
        by_issue[r["issue"]] = by_issue.get(r["issue"], 0) + 1
    elapsed = time.monotonic() - t0
    metrics = {"n_rows": float(n), "n_flagged": float(len(flagged)),
               "flag_rate_pct": round(100.0 * len(flagged) / n, 3),
               "elapsed_seconds": round(elapsed, 3)}
    for issue, count in by_issue.items():
        metrics[f"n_{issue}"] = float(count)
    if out_rows and "sme_rating" in out_rows[0]:
        fp = sum(1 for r in flagged if r["sme_rating"] in ACCEPT_RATINGS)
        tp = sum(1 for r in flagged if r["sme_rating"] in REJECT_RATINGS)
        metrics["flagged_sme_acceptable"] = float(fp)
        metrics["flagged_sme_nonacceptable"] = float(tp)
        metrics["flag_precision_pct"] = (
            round(100.0 * tp / len(flagged), 3) if flagged else 0.0)
    msg = (f"flagged {len(flagged)}/{n} contrast-fidelity issue(s) "
           + (f"({', '.join(f'{k}={v}' for k, v in sorted(by_issue.items()))})"
              if by_issue else "(clean)"))
    return FunctionResult(ok=True, outputs={"flags": str(out)},
                          metrics=metrics, message=msg)


# ---------------------------------------------------------------------------
# sme_metric_separation
# ---------------------------------------------------------------------------


def sme_metric_separation(ctx: RunContext, inputs: dict[str, Any],
                          params: dict[str, Any]) -> FunctionResult:
    """How well does each candidate metric reproduce the SME's verdicts?"""
    lpath = _dataset_path(inputs.get("labels"))
    if not lpath or not Path(lpath).exists():
        return FunctionResult(
            ok=False, message="sme_metric_separation: no `labels` dataset wired")
    cand_col = str(params.get("candidate_col") or "reviewed_ko")
    ref_col = str(params.get("reference_col") or "ko_reference")
    allrefs_col = str(params.get("allrefs_col") or "ko_all")
    label_col = str(params.get("label_col") or "sme_rating")
    sep = str(params.get("multi_ref_separator") or "|")

    rows: list[dict] = []
    with Path(lpath).open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            cand = (r.get(cand_col) or "").strip()
            refs_raw = (r.get(allrefs_col) or r.get(ref_col) or "").strip()
            refs = [x for x in refs_raw.split(sep) if x.strip()]
            rating = (r.get(label_col) or "").strip().upper()
            if not cand or not refs or rating not in ACCEPT_RATINGS | REJECT_RATINGS:
                continue
            rows.append({"sctid": (r.get("sctid") or "").strip(),
                         "candidate": cand, "refs": refs, "rating": rating})
    if not rows:
        return FunctionResult(ok=False,
                              message=f"sme_metric_separation: no usable rows in {lpath}")

    labels = [int(r["rating"] in ACCEPT_RATINGS) for r in rows]

    exact_scores = [
        float(any(norm_text(r["candidate"]) == norm_text(ref) for ref in r["refs"]))
        for r in rows]
    chrf_scores = [
        max(sacrebleu.sentence_chrf(r["candidate"], [ref]).score
            for ref in r["refs"]) / 100.0
        for r in rows]

    # BGE-M3 cosine vs each ref, keep the max — one batched call per side.
    pairs = [(i, ref) for i, r in enumerate(rows) for ref in r["refs"]]
    sims = _embed_similarity([rows[i]["candidate"] for i, _ in pairs],
                             [ref for _, ref in pairs])
    cos_scores = [0.0] * len(rows)
    for (i, _), sim in zip(pairs, sims):
        cos_scores[i] = max(cos_scores[i], sim)

    def class_means(scores: list[float]) -> tuple[float, float]:
        pos = [s for s, y in zip(scores, labels) if y == 1]
        neg = [s for s, y in zip(scores, labels) if y == 0]
        return (sum(pos) / len(pos) if pos else float("nan"),
                sum(neg) / len(neg) if neg else float("nan"))

    metrics: dict[str, float] = {"n_rows": float(len(rows)),
                                 "n_acceptable": float(sum(labels)),
                                 "n_not_acceptable": float(len(labels) - sum(labels))}
    audit_cols: dict[str, list[float]] = {}
    for name, scores in (("exact", exact_scores), ("chrf", chrf_scores),
                         ("cosine", cos_scores)):
        pos_mean, neg_mean = class_means(scores)
        metrics[f"{name}_auc"] = round(_auc(scores, labels), 4)
        metrics[f"{name}_mean_acceptable"] = round(pos_mean, 4)
        metrics[f"{name}_mean_not_acceptable"] = round(neg_mean, 4)
        audit_cols[name] = scores
    thr, thr_stats = _best_threshold(cos_scores, labels)
    metrics["cosine_best_threshold"] = round(float(thr), 4)
    metrics["cosine_best_threshold_balanced_acc"] = round(
        float(thr_stats["balanced"]), 4)

    # The honest view: identity pairs (candidate string == a gold ref) make
    # separation trivial, because for Correct rows the gold IS the reviewed
    # translation. Restrict to rows where the candidate differs from every
    # reference as a raw string — positives there are exactly the "acceptable
    # variation" (spacing / preferred-synonym) the SME wants accepted, and a
    # metric only earns its keep by separating those from Partial/Wrong.
    nonident = [i for i, r in enumerate(rows)
                if all(r["candidate"] != ref for ref in r["refs"])]
    metrics["n_nonidentical"] = float(len(nonident))
    ni_labels = [labels[i] for i in nonident]
    metrics["n_nonidentical_acceptable"] = float(sum(ni_labels))
    if ni_labels and 0 < sum(ni_labels) < len(ni_labels):
        for name, scores in (("exact", exact_scores), ("chrf", chrf_scores),
                             ("cosine", cos_scores)):
            ni_scores = [scores[i] for i in nonident]
            metrics[f"{name}_auc_nonidentical"] = round(
                _auc(ni_scores, ni_labels), 4)
        ni_thr, ni_stats = _best_threshold(
            [cos_scores[i] for i in nonident], ni_labels)
        metrics["cosine_nonidentical_threshold"] = round(float(ni_thr), 4)
        metrics["cosine_nonidentical_balanced_acc"] = round(
            float(ni_stats["balanced"]), 4)

    out = Path(ctx.log_dir) / "sme_metric_separation_audit.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sctid", "candidate", "canonical_ref", "n_refs", "rating",
                    "exact", "chrf", "cosine"])
        for i, r in enumerate(rows):
            w.writerow([r["sctid"], r["candidate"], r["refs"][0], len(r["refs"]),
                        r["rating"], exact_scores[i], round(chrf_scores[i], 4),
                        round(cos_scores[i], 4)])

    msg = (f"{len(rows)} rows: AUC exact={metrics['exact_auc']} "
           f"chrf={metrics['chrf_auc']} cosine={metrics['cosine_auc']} "
           f"(cosine thr {metrics['cosine_best_threshold']} -> "
           f"balanced acc {metrics['cosine_best_threshold_balanced_acc']})")
    return FunctionResult(ok=True, outputs={"audit": str(out)},
                          metrics=metrics, message=msg)


# ---------------------------------------------------------------------------
# self_review — can a model catch its own mistakes?
# ---------------------------------------------------------------------------

# Deliberately NEUTRAL. It names no error class (no mention of contrast,
# suffixes, register, word order, dropped modifiers) and gives no style guide,
# so a flagged row reflects the model's own judgement rather than a checklist
# we handed it. Telling it what to look for would test instruction-following,
# not self-critique.
SELF_REVIEW_SYSTEM = """You are reviewing Korean translations of English SNOMED CT clinical terms.

For each pair you are given the English source term and a proposed Korean translation.
Decide whether the proposed translation should be used as-is, or changed.

Reply with STRICT JSON only, no prose around it:
{"verdict": "KEEP|CHANGE", "revision": "<the Korean you would use instead, or empty if KEEP>", "reason": "<one short sentence>"}

If you would use the proposed translation unchanged, reply KEEP with an empty revision."""


def self_review(ctx: RunContext, inputs: dict[str, Any],
                params: dict[str, Any]) -> FunctionResult:
    """Ask a model to review translations (usually its own) and measure it.

    Reports, against the SME gold when a ``gold`` dataset is wired:
      detection  — CHANGE rate on rows that did NOT match gold (mistakes found)
      false_alarm— CHANGE rate on rows that DID match gold (needless edits)
      repair     — of the CHANGEd wrong rows, how many revisions reach gold
      damage     — of the CHANGEd correct rows, how many revisions leave gold
    """
    import json as _json
    from concurrent.futures import ThreadPoolExecutor

    from snomed_translation.acceptability import _judge_claude, _judge_local, _is_claude

    tpath = _dataset_path(inputs.get("translations"))
    if not tpath or not Path(tpath).exists():
        return FunctionResult(ok=False, message="self_review: no `translations` wired")
    roles = _roles(inputs.get("translations"))
    id_col = _col(params, roles, "id_col", "sctid", "sctid")
    en_col = _col(params, roles, "en_col", "en", "preferred_term")
    ko_col = _col(params, roles, "ko_col", "target", "translation")
    model = str(params.get("model") or "")
    if not model:
        return FunctionResult(ok=False, message="self_review: `model` param required")
    base_url = str(params.get("base_url") or "http://localhost:8086")
    system = str(params.get("system") or SELF_REVIEW_SYSTEM)
    # Optional: give the reviewer the same style guide the translator had. The
    # unguided arm measures the model's own prior; the guided arm measures
    # whether it can apply a rulebook it demonstrably follows when translating.
    sg = inputs.get("style_guide")
    sg_text = ""
    if isinstance(sg, dict):
        sg_text = sg.get("text") or sg.get("body") or ""
        if not sg_text and sg.get("path") and Path(sg["path"]).exists():
            sg_text = Path(sg["path"]).read_text(encoding="utf-8")
    elif isinstance(sg, str) and sg.strip():
        sg_text = (Path(sg).read_text(encoding="utf-8")
                   if Path(sg).exists() else sg)
    if sg_text:
        system = f"{system}\n\n# Style guide\n\n{sg_text}"
    max_tokens = int(params.get("max_tokens") or 220)
    concurrency = int(params.get("concurrency") or (4 if _is_claude(model) else 16))

    gold: dict[str, set[str]] = {}
    gpath = _dataset_path(inputs.get("gold"))
    if gpath and Path(gpath).exists():
        allrefs = str(params.get("allrefs_col") or "ko_all")
        refcol = str(params.get("reference_col") or "ko_reference")
        with Path(gpath).open(encoding="utf-8") as f:
            for r in csv.DictReader(f):
                raw = (r.get(allrefs) or r.get(refcol) or "")
                gold[(r.get("sctid") or "").strip()] = {
                    norm_text(x) for x in raw.split("|") if x.strip()}

    with Path(tpath).open(encoding="utf-8") as f:
        rows = [r for r in csv.DictReader(f)
                if (r.get(en_col) or "").strip() and (r.get(ko_col) or "").strip()]

    def one(r: dict) -> dict:
        en, ko = r[en_col].strip(), r[ko_col].strip()
        payload = f"english: {en}\nkorean: {ko}"
        try:
            if _is_claude(model):
                from snomed_translation.generate import run_query
                text = run_query(payload, model=model, system=system,
                                 thinking=False, ctx=ctx)
            else:
                import urllib.request
                body = _json.dumps({
                    "model": model, "temperature": 0.0, "max_tokens": max_tokens,
                    "messages": [{"role": "system", "content": system},
                                 {"role": "user", "content": payload}],
                }).encode()
                req = urllib.request.Request(
                    base_url.rstrip("/") + "/v1/chat/completions", data=body,
                    headers={"Content-Type": "application/json"})
                resp = _json.loads(urllib.request.urlopen(req, timeout=180).read())
                from pipelines.llm_accounting import record_completion
                record_completion(ctx, model=model, usage=resp.get("usage"))
                text = resp["choices"][0]["message"]["content"]
            m = re.search(r"\{.*\}", text, re.S)
            d = _json.loads(m.group(0)) if m else {}
        except Exception as exc:  # a failed review must not kill the run
            d = {"verdict": "ERROR", "revision": "", "reason": str(exc)[:120]}
        verdict = str(d.get("verdict") or "").strip().upper()
        revision = str(d.get("revision") or "").strip()
        sid = (r.get(id_col) or "").strip()
        refs = gold.get(sid)
        was_right = None if refs is None else int(norm_text(ko) in refs)
        rev_right = None
        if refs is not None and verdict == "CHANGE" and revision:
            rev_right = int(norm_text(revision) in refs)
        return {"sctid": sid, "english": en, "korean": ko, "verdict": verdict,
                "revision": revision, "reason": str(d.get("reason") or "")[:200],
                "was_correct": "" if was_right is None else was_right,
                "revision_correct": "" if rev_right is None else rev_right}

    with ThreadPoolExecutor(max_workers=max(1, concurrency)) as ex:
        out_rows = list(ex.map(one, rows))

    out = Path(ctx.log_dir) / "self_review.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        w.writeheader()
        w.writerows(out_rows)

    n = len(out_rows)
    changed = [r for r in out_rows if r["verdict"] == "CHANGE"]
    metrics = {"n_rows": float(n), "n_change": float(len(changed)),
               "n_keep": float(sum(1 for r in out_rows if r["verdict"] == "KEEP")),
               "n_error": float(sum(1 for r in out_rows if r["verdict"] == "ERROR")),
               "change_rate_pct": round(100.0 * len(changed) / n, 2)}
    scored = [r for r in out_rows if r["was_correct"] != ""]
    if scored:
        wrong = [r for r in scored if r["was_correct"] == 0]
        right = [r for r in scored if r["was_correct"] == 1]
        ch_wrong = [r for r in wrong if r["verdict"] == "CHANGE"]
        ch_right = [r for r in right if r["verdict"] == "CHANGE"]
        metrics["n_was_wrong"] = float(len(wrong))
        metrics["n_was_correct"] = float(len(right))
        metrics["detection_rate_pct"] = round(
            100.0 * len(ch_wrong) / len(wrong), 2) if wrong else 0.0
        metrics["false_alarm_rate_pct"] = round(
            100.0 * len(ch_right) / len(right), 2) if right else 0.0
        metrics["repair_rate_pct"] = round(
            100.0 * sum(1 for r in ch_wrong if r["revision_correct"] == 1)
            / len(ch_wrong), 2) if ch_wrong else 0.0
        metrics["damage_rate_pct"] = round(
            100.0 * sum(1 for r in ch_right if r["revision_correct"] == 0)
            / len(ch_right), 2) if ch_right else 0.0
        net = (sum(1 for r in ch_wrong if r["revision_correct"] == 1)
               - sum(1 for r in ch_right if r["revision_correct"] == 0))
        metrics["net_rows_gained"] = float(net)

    msg = (f"{len(changed)}/{n} CHANGE"
           + (f"; detection {metrics.get('detection_rate_pct')}% "
              f"false-alarm {metrics.get('false_alarm_rate_pct')}% "
              f"repair {metrics.get('repair_rate_pct')}% "
              f"damage {metrics.get('damage_rate_pct')}% "
              f"net {int(metrics.get('net_rows_gained', 0)):+d} rows" if scored else ""))
    return FunctionResult(ok=True, outputs={"reviews": str(out)},
                          metrics=metrics, message=msg)


# ---------------------------------------------------------------------------
# escalate_uncertain — confidence-routed two-model cascade
# ---------------------------------------------------------------------------

ESCALATE_SUFFIX = """
A smaller model produced these candidate translations for this term (with how
many of its samples gave each). It was NOT confident — the samples disagreed.
Use them as evidence of where the difficulty lies, not as options to pick from:
you may output one of them, a modification, or something different.

{candidates}

Output ONLY your final Korean translation."""


def escalate_uncertain(ctx: RunContext, inputs: dict[str, Any],
                       params: dict[str, Any]) -> FunctionResult:
    """Keep the confident model's answer; re-translate the uncertain rows.

    Routing uses sampling disagreement (``n_distinct``), the strongest
    confidence signal measured on this task (AUC 0.755 for predicting
    incorrectness). Rows below the threshold keep the sampler's unanimous
    answer; rows at or above it are re-translated by a second, stronger model
    which REPLAYS THE ORIGINAL PROMPT (same style guide, same retrieved
    exemplars) and — when ``show_candidates`` — additionally sees what the
    first model produced. It generates rather than selects, because selection
    among the first model's samples was measured to gain nothing.
    """
    import json as _json
    import urllib.request
    from concurrent.futures import ThreadPoolExecutor

    cpath = _dataset_path(inputs.get("candidates"))
    if not cpath or not Path(cpath).exists():
        return FunctionResult(ok=False, message="escalate_uncertain: no `candidates` wired")
    side = Path(str(cpath).replace(".csv", ".prompts.json"))
    if not side.exists():
        return FunctionResult(
            ok=False, message=f"escalate_uncertain: prompt sidecar missing at {side}")
    sidecar = _json.loads(side.read_text(encoding="utf-8"))
    system = sidecar.get("system_prompt") or sidecar.get("system") or ""
    user_prompts = sidecar.get("user_prompts") or sidecar.get("prompts") or {}

    model = str(params.get("model") or "")
    base_url = str(params.get("base_url") or "")
    if not model or not base_url:
        return FunctionResult(
            ok=False, message="escalate_uncertain: `model` and `base_url` required")
    api_key = os.environ.get(str(params.get("api_key_env") or ""), "")
    min_distinct = int(params.get("min_distinct") or 2)
    show_candidates = bool(params.get("show_candidates", True))
    max_escalate = int(params.get("max_escalate") or 0)
    concurrency = int(params.get("concurrency") or 8)
    llm = dict(params.get("llm_params") or
               {"temperature": 0.0, "max_tokens": 256, "enable_thinking": False})

    with Path(cpath).open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    normalize_spacing = bool(params.get("normalize_spacing", True))

    def n_distinct(r) -> int:
        """Distinct-candidate count driving the routing decision.

        With ``normalize_spacing`` (default) the vote ignores 띄어쓰기, which
        the SME has ruled is never itself an error. Otherwise spacing-only
        variation escalates rows that need no escalation: 5 of 40 assessed
        escalations were byte-identical no-ops, and some churn came from
        re-translating an answer that was already gold modulo spacing.
        """
        if normalize_spacing:
            try:
                cands = _json.loads(r.get("candidates") or "[]")
            except Exception:
                cands = []
            forms = {norm_text(c.get("text", "")) for c in cands
                     if (c.get("text") or "").strip()}
            if forms:
                return len(forms)
        try:
            return int(r.get("n_distinct") or 1)
        except ValueError:
            return 1

    uncertain = sorted([r for r in rows if n_distinct(r) >= min_distinct],
                       key=lambda r: -n_distinct(r))
    if max_escalate:
        uncertain = uncertain[:max_escalate]
    esc_ids = {r["sctid"] for r in uncertain}

    n_total = len(uncertain)
    n_done = [0]
    t0 = time.monotonic()
    progress_lock = Lock()

    def call(r: dict) -> tuple[str, str]:
        sid = r["sctid"]
        user = user_prompts.get(sid, "")
        if show_candidates:
            try:
                cands = _json.loads(r.get("candidates") or "[]")
            except Exception:
                cands = []
            block = "\n".join(
                f"- {c.get('text','')}  ({c.get('count',1)} of "
                f"{r.get('n_samples','?')} samples)" for c in cands)
            user = user.rstrip() + "\n" + ESCALATE_SUFFIX.format(candidates=block)
        body = _json.dumps({"model": model, "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user}], **llm}).encode()
        headers = {"Content-Type": "application/json"}
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"
        # Retry: a failed escalation silently falls back to the LOW-CONFIDENCE
        # kept answer (a 15-18%-correct stratum), so a transient network blip
        # costs real accuracy. 1/103 failed in a 200-row run; at ~2,500
        # production calls that is ~25 rows.
        attempts = int(params.get("max_attempts") or 3)
        last = ""
        for attempt in range(1, attempts + 1):
            try:
                req = urllib.request.Request(
                    base_url.rstrip("/") + "/v1/chat/completions", data=body,
                    headers=headers)
                resp = _json.loads(urllib.request.urlopen(req, timeout=180).read())
                from pipelines.llm_accounting import record_completion
                record_completion(ctx, model=model, usage=resp.get("usage"))
                text = resp["choices"][0]["message"]["content"].strip()
                _progress()
                return sid, text
            except Exception as exc:
                last = str(exc)
                if attempt < attempts:
                    time.sleep(min(2 ** attempt, 8))
        _progress()
        return sid, f"ERROR: {last}"

    def _progress() -> None:
        with progress_lock:
            n_done[0] += 1
            done = n_done[0]
        if done % 100 == 0 or done == n_total:
            elapsed = time.monotonic() - t0
            rate = done / elapsed if elapsed else 0.0
            eta = (n_total - done) / rate if rate else 0.0
            log.info("[escalate_uncertain] %d/%d escalations (%.1f req/s, "
                     "ETA %.0fs)", done, n_total, rate, eta)

    escalated: dict[str, str] = {}
    if uncertain:
        log.info("[escalate_uncertain] escalating %d/%d rows (n_distinct>=%d) "
                 "to %s, concurrency=%d", n_total, len(rows), min_distinct,
                 model, concurrency)
        from snomed_translation.watchdog import progress_watchdog
        with progress_watchdog("escalate_uncertain", stall_seconds=180.0), \
             ThreadPoolExecutor(max_workers=max(1, concurrency)) as ex:
            escalated = dict(ex.map(call, uncertain))

    # Contrast gate: reject a revision that INTRODUCES a contrast-fidelity
    # fault the kept answer did not have. Inventing 조영제 사용 is the error
    # class the SME flagged most, and an escalation was observed doing exactly
    # that (241183005). The detector has zero measured false positives, so
    # this is a safe veto — the row simply keeps the first model's answer.
    gate_contrast = bool(params.get("gate_contrast", True))
    n_gated = 0
    if gate_contrast:
        by_id = {r["sctid"]: r for r in rows}
        for sid, text in list(escalated.items()):
            if text.startswith("ERROR"):
                continue
            src = by_id.get(sid, {})
            en = src.get("preferred_term", "")
            kept_text = src.get("top_candidate", "")
            new_issue = contrast_issue(en, text)
            if new_issue in ("hallucinated", "wrong_polarity") and \
                    contrast_issue(en, kept_text) != new_issue:
                escalated.pop(sid)
                esc_ids.discard(sid)
                n_gated += 1

    out = Path(ctx.artifacts_dir() or ctx.log_dir) / (
        f"cascade_{params.get('output_tag') or 'run'}.csv")
    out.parent.mkdir(parents=True, exist_ok=True)
    n_err = 0
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "sctid", "preferred_term", "ko_reference", "translation",
            "n_distinct", "routed"])
        w.writeheader()
        for r in rows:
            sid = r["sctid"]
            if sid in escalated and not escalated[sid].startswith("ERROR"):
                text, routed = escalated[sid], "escalated"
            elif sid in escalated:
                text, routed, n_err = r.get("top_candidate", ""), "escalation_failed", n_err + 1
            else:
                text, routed = r.get("top_candidate", ""), "kept"
            w.writerow({"sctid": sid, "preferred_term": r.get("preferred_term", ""),
                        "ko_reference": r.get("ko_reference", ""),
                        "translation": text, "n_distinct": n_distinct(r),
                        "routed": routed})

    metrics = {"n_rows": float(len(rows)), "n_escalated": float(len(esc_ids)),
               "n_gated_contrast": float(n_gated),
               "normalize_spacing": float(bool(normalize_spacing)),
               "n_kept": float(len(rows) - len(esc_ids)),
               "escalation_rate_pct": round(100.0 * len(esc_ids) / len(rows), 2),
               "n_escalation_errors": float(n_err),
               "min_distinct": float(min_distinct),
               "show_candidates": float(bool(show_candidates))}
    msg = (f"kept {len(rows)-len(esc_ids)}, escalated {len(esc_ids)} "
           f"(n_distinct>={min_distinct}) to {model}"
           + (" WITH candidates shown" if show_candidates else " blind")
           + (f"; {n_gated} revisions vetoed by the contrast gate" if n_gated else "")
           + (f"; {n_err} escalation errors" if n_err else ""))
    return FunctionResult(ok=True, outputs={"translations": str(out)},
                          metrics=metrics, message=msg)


# ---------------------------------------------------------------------------
# ingest_review_pack / priority_tier_separation (review round 3, 2026-08)
# ---------------------------------------------------------------------------

# The reviewer-facing workbook layout, as package_deliverable writes it. The
# returned file's header row cannot be trusted verbatim: round 3 came back
# with `sme_error_category` overwritten by a second `sme_corrected_ko`, which
# a name-keyed reader silently collapses (DictReader keeps the LAST duplicate,
# so every correction would have been read as a category and dropped).
_REVIEW_SHEET_CANONICAL = [
    "sctid", "preferred_term", "translation_ko", "synonyms_ko",
    "sme_rating", "sme_corrected_ko", "sme_error_category", "sme_notes",
]

_RATINGS = ("ACCEPTABLE", "PARTIAL", "WRONG", "CORRECT")


def ingest_review_pack(ctx: RunContext, inputs: dict[str, Any],
                       params: dict[str, Any]) -> FunctionResult:
    """Parse a returned SME review workbook into a normalised dataset.

    Reads one sheet of the .xlsx the reviewer sent back, repairs a damaged
    header row by position against the canonical layout, normalises ratings,
    and tallies the reviewer's own error categories. The output CSV is the
    single source row set for every downstream analysis of the round, so the
    round is ingested exactly once, by a tracked run, not re-parsed ad hoc.
    """
    from openpyxl import load_workbook

    xlsx = str(params.get("xlsx_path") or "").strip()
    if not xlsx or not Path(xlsx).exists():
        return FunctionResult(ok=False, message=f"ingest_review_pack: no workbook at {xlsx!r}")
    sheet = str(params.get("sheet") or "Sample - please do these first")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return FunctionResult(ok=False,
                              message=f"ingest_review_pack: sheet {sheet!r} not in {wb.sheetnames}")
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    if not rows:
        return FunctionResult(ok=False, message="ingest_review_pack: empty sheet")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_repaired = 0
    if len(set(header)) != len(header) or header != _REVIEW_SHEET_CANONICAL[:len(header)]:
        if len(header) == len(_REVIEW_SHEET_CANONICAL):
            header = list(_REVIEW_SHEET_CANONICAL)
            header_repaired = 1
        else:
            return FunctionResult(
                ok=False,
                message=(f"ingest_review_pack: header {header} has duplicates or "
                         f"unexpected names and column count {len(header)} != "
                         f"{len(_REVIEW_SHEET_CANONICAL)}, cannot repair by position"))

    out_rows: list[dict[str, str]] = []
    counts = {r: 0 for r in _RATINGS}
    n_unrated = n_corrected = n_categorized = 0
    cats: dict[str, int] = {}
    for raw in rows[1:]:
        raw = list(raw) + [None] * (len(header) - len(raw))
        rec = {h: (str(v).strip() if v is not None else "") for h, v in zip(header, raw)}
        if not rec.get("sctid"):
            continue
        rating = rec.get("sme_rating", "").upper()
        rec["sme_rating"] = rating
        if rating in counts:
            counts[rating] += 1
        elif rating:
            return FunctionResult(ok=False,
                                  message=f"ingest_review_pack: unknown rating {rating!r} "
                                          f"on {rec['sctid']}")
        else:
            n_unrated += 1
        if rec.get("sme_corrected_ko"):
            n_corrected += 1
        if rec.get("sme_error_category"):
            n_categorized += 1
            for c in rec["sme_error_category"].replace(";", ",").split(","):
                c = re.sub(r"\s+", " ", c.strip().lower())
                if c:
                    cats[c] = cats.get(c, 0) + 1
        out_rows.append(rec)

    tag = str(params.get("output_tag") or "").strip()
    out = Path(ctx.log_dir) / f"sme_review_ingested{'_' + tag if tag else ''}.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(header))
        w.writeheader()
        w.writerows(out_rows)

    # Terms output: the reviewed rows reshaped as an eval/translate set. The
    # effective reference is the reviewer's correction when one was given,
    # otherwise the text she rated (ACCEPTABLE rows are their own reference).
    # A correction may span multiple lines (preferred term first, synonyms
    # after, sometimes labelled "(as a synonym)"); the reference is the first
    # line. `typo_fixes` ("wrong=right;...") applies ONLY here — the reviewed
    # output above keeps her text verbatim.
    fixes = [p.split("=", 1) for p in
             str(params.get("typo_fixes") or "").split(";") if "=" in p]
    n_typo_fixed = 0
    terms_rows = []
    for rec in out_rows:
        ref = (rec.get("sme_corrected_ko") or "").splitlines()[0].strip() \
            if rec.get("sme_corrected_ko") else rec.get("translation_ko", "")
        ref = ref.strip().strip('"')
        for wrong, right in fixes:
            if wrong in ref:
                ref = ref.replace(wrong, right)
                n_typo_fixed += 1
        terms_rows.append({"sctid": rec["sctid"],
                           "preferred_term": rec["preferred_term"],
                           "ko_reference": ref,
                           "sme_rating": rec["sme_rating"]})
    terms_out = Path(ctx.log_dir) / f"sme_review_terms{'_' + tag if tag else ''}.csv"
    with terms_out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["sctid", "preferred_term",
                                          "ko_reference", "sme_rating"])
        w.writeheader()
        w.writerows(terms_rows)

    metrics: dict[str, float] = {
        "n_rows": float(len(out_rows)),
        "n_rated": float(sum(counts.values())),
        "n_unrated": float(n_unrated),
        "n_corrected": float(n_corrected),
        "n_categorized": float(n_categorized),
        "header_repaired": float(header_repaired),
        "n_typo_fixed": float(n_typo_fixed),
    }
    for r, c in counts.items():
        metrics[f"n_{r.lower()}"] = float(c)
    for c, k in sorted(cats.items(), key=lambda kv: -kv[1]):
        metrics["cat_" + re.sub(r"[^a-z0-9]+", "_", c).strip("_")] = float(k)

    msg = (f"{len(out_rows)} rows, {sum(counts.values())} rated ("
           + ", ".join(f"{v} {k}" for k, v in counts.items() if v)
           + f"), {n_corrected} corrected"
           + ("; header row repaired by position" if header_repaired else "")
           + (f"; {n_typo_fixed} typo fix(es) in terms refs" if n_typo_fixed else ""))
    return FunctionResult(ok=True, outputs={"reviewed": str(out),
                                            "terms": str(terms_out)},
                          metrics=metrics, message=msg)


def priority_tier_separation(ctx: RunContext, inputs: dict[str, Any],
                             params: dict[str, Any]) -> FunctionResult:
    """Does review_priority predict what the SME actually rejects?

    Joins a reviewed sample to the pack that carries the (blinded) tier
    labels and tests whether rejection rises across ordered tiers with the
    linear-by-linear association test — the pre-registered analysis for the
    round-3 sample: it equals Cochran-Armitage for a binary outcome and keeps
    the ACCEPTABLE < PARTIAL < WRONG ordering (both recover power that
    collapsing to right/wrong throws away).
    """
    import math

    rpath = _dataset_path(inputs.get("reviewed"))
    ppath = _dataset_path(inputs.get("pack"))
    if not rpath or not Path(rpath).exists():
        return FunctionResult(ok=False, message="priority_tier_separation: no `reviewed`")
    if not ppath or not Path(ppath).exists():
        return FunctionResult(ok=False, message="priority_tier_separation: no `pack`")

    tier_col = str(params.get("tier_col") or "review_priority")
    rating_col = str(params.get("rating_col") or "sme_rating")

    tiers: dict[str, str] = {}
    with Path(ppath).open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            sid = (r.get("sctid") or "").strip()
            if sid:
                tiers[sid] = (r.get(tier_col) or "").strip().lower()

    tier_score = {"low": 0.0, "medium": 1.0, "high": 2.0}
    rating_score = {"ACCEPTABLE": 0.0, "CORRECT": 0.0, "PARTIAL": 1.0, "WRONG": 2.0}

    joined: list[dict[str, str]] = []
    n_unmatched = 0
    with Path(rpath).open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            sid = (r.get("sctid") or "").strip()
            rating = (r.get(rating_col) or "").strip().upper()
            tier = tiers.get(sid, "")
            if not rating:
                continue
            if tier not in tier_score:
                n_unmatched += 1
                continue
            joined.append({"sctid": sid, "tier": tier, "rating": rating,
                           "preferred_term": (r.get("preferred_term") or "").strip(),
                           "sme_error_category": (r.get("sme_error_category") or "").strip()})

    if not joined:
        return FunctionResult(ok=False, message="priority_tier_separation: nothing joined")

    def linear_by_linear(xs: list[float], ys: list[float]) -> tuple[float, float]:
        n = len(xs)
        mx, my = sum(xs) / n, sum(ys) / n
        sxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
        sxx = sum((x - mx) ** 2 for x in xs)
        syy = sum((y - my) ** 2 for y in ys)
        if sxx <= 0 or syy <= 0:
            return 0.0, 1.0
        r = sxy / math.sqrt(sxx * syy)
        z = r * math.sqrt(n - 1)
        return z, math.erfc(abs(z) / math.sqrt(2.0))  # two-sided

    xs = [tier_score[j["tier"]] for j in joined]
    ys_bin = [1.0 if rating_score[j["rating"]] > 0 else 0.0 for j in joined]
    ys_ord = [rating_score[j["rating"]] for j in joined]
    z_bin, p_bin = linear_by_linear(xs, ys_bin)
    z_ord, p_ord = linear_by_linear(xs, ys_ord)

    metrics: dict[str, float] = {
        "n_joined": float(len(joined)), "n_unmatched": float(n_unmatched),
        "trend_z_binary": round(z_bin, 4), "trend_p_binary": round(p_bin, 6),
        "trend_z_ordinal": round(z_ord, 4), "trend_p_ordinal": round(p_ord, 6),
    }
    for tier in ("low", "medium", "high"):
        rows_t = [j for j in joined if j["tier"] == tier]
        n_rej = sum(1 for j in rows_t if rating_score[j["rating"]] > 0)
        n_wrong = sum(1 for j in rows_t if j["rating"] == "WRONG")
        metrics[f"{tier}_n"] = float(len(rows_t))
        metrics[f"{tier}_n_not_acceptable"] = float(n_rej)
        metrics[f"{tier}_n_wrong"] = float(n_wrong)
        metrics[f"{tier}_reject_rate_pct"] = (
            round(100.0 * n_rej / len(rows_t), 2) if rows_t else 0.0)

    tag = str(params.get("output_tag") or "").strip()
    out = Path(ctx.log_dir) / f"tier_ratings{'_' + tag if tag else ''}.csv"
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["sctid", "tier", "rating",
                                          "preferred_term", "sme_error_category"])
        w.writeheader()
        w.writerows(joined)

    msg = ("reject rate " + " / ".join(
        f"{t}={metrics[f'{t}_reject_rate_pct']:.0f}% (n={int(metrics[f'{t}_n'])})"
        for t in ("low", "medium", "high"))
        + f"; trend p={p_bin:.4g} binary, p={p_ord:.4g} ordinal")
    return FunctionResult(ok=True, outputs={"joined": str(out)},
                          metrics=metrics, message=msg)


def merge_adjudicated_gold(ctx: RunContext, inputs: dict[str, Any],
                           params: dict[str, Any]) -> FunctionResult:
    """Merge a new review round into the adjudicated gold set, with supersession.

    The adjudicated set is the SME lock and the reviewer overlay, so what it
    contains IS the system's definition of "currently ruled". Two laws:

    1. RECENCY WINS. A round-3 row replaces an older gold row for the same
       concept, and an older gold row whose text violates the CURRENT rule
       file (which encodes the newest rulings) is withheld as superseded —
       emitted on the ``superseded`` output as the reviewer's confirmation
       list, never silently dropped. Without this, an immutable lock preserves
       text its own author has overruled: 20 such rows after round 3.
    2. NEWEST ROWS ARE NEVER RULE-CHECKED AWAY. The latest round is
       definitionally current; where it contradicts a rule derived from an
       OLDER ruling (round 3's arthrogram 조영상 vs the batch-2 조영상 ban),
       the contradiction goes to the adjudication ledger, not auto-resolution.

    Round-3 semantics: ko_reference = the reviewer's correction (first line,
    typo-fixed) where given, else the text she rated. A correction's later
    lines are synonyms ONLY when explicitly annotated "(… synonym)" — other
    trailing commentary is not an endorsement. An ACCEPTABLE row WITH a
    correction keeps the machine text as an accepted synonym; a PARTIAL row
    does not (she rated it partial, it is not an accepted form).
    """
    from snomed_translation.hard_rules import find_violations, load_hard_rules

    gold_rows = []
    gpath = _dataset_path(inputs.get("gold"))
    if not gpath or not Path(gpath).exists():
        return FunctionResult(ok=False, message="merge_adjudicated_gold: no `gold`")
    with Path(gpath).open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fields = list(reader.fieldnames or [])
        gold_rows = list(reader)

    rpath = _dataset_path(inputs.get("round"))
    if not rpath or not Path(rpath).exists():
        return FunctionResult(ok=False, message="merge_adjudicated_gold: no `round`")
    with Path(rpath).open(encoding="utf-8") as f:
        round_rows = list(csv.DictReader(f))

    severities = {s.strip().lower() for s in
                  str(params.get("severities") or "blocker,warning").split(",")
                  if s.strip()}
    rules = [r for r in load_hard_rules(params.get("rules_file"))
             if r.severity in severities] if params.get("rules_file") else []
    fixes = [p.split("=", 1) for p in
             str(params.get("typo_fixes") or "").split(";") if "=" in p]
    batch_label = str(params.get("batch_label") or "round3")
    syn_re = re.compile(r"\(([^)]*synonym[^)]*)\)\s*$", re.I)

    new_ids = {(r.get("sctid") or "").strip() for r in round_rows}
    merged: list[dict] = []
    superseded: list[dict] = []
    per_check: dict[str, int] = {}
    n_replaced = 0
    for g in gold_rows:
        sid = (g.get("sctid") or "").strip()
        if sid in new_ids:
            n_replaced += 1
            superseded.append({**g, "superseded_by": "replaced_by_" + batch_label})
            continue
        found = find_violations((g.get("ko_reference") or "").strip(), rules,
                                require_enforce=False,
                                source=(g.get("preferred_term") or "").strip())
        if found:
            checks = ",".join(sorted(r.id for r, _ in found))
            for r, _ in found:
                per_check[r.id] = per_check.get(r.id, 0) + 1
            superseded.append({**g, "superseded_by": checks})
            continue
        merged.append(g)

    n_round_corrected = 0
    for r in round_rows:
        machine = (r.get("translation_ko") or "").strip()
        corr_raw = (r.get("sme_corrected_ko") or "").strip()
        lines = [l.strip().strip('"') for l in corr_raw.splitlines() if l.strip()]
        text = lines[0] if lines else machine
        for wrong, right in fixes:
            text = text.replace(wrong, right)
        syns = [syn_re.sub("", l).strip() for l in lines[1:] if syn_re.search(l)]
        rating = (r.get("sme_rating") or "").strip().upper()
        flags = ""
        if lines:
            n_round_corrected += 1
            if rating == "ACCEPTABLE" and machine and machine != text:
                syns.append(machine)
                flags = "sme_original_as_synonym"
        merged.append({
            "sctid": (r.get("sctid") or "").strip(),
            "preferred_term": (r.get("preferred_term") or "").strip(),
            "ko_reference": text,
            "ko_all": "|".join([text] + [s for s in syns if s]),
            "reviewed_ko": text if lines else machine,
            "batch": batch_label,
            "sme_rating": rating,
            "gold_flags": flags,
        })

    tag = str(params.get("output_tag") or batch_label)
    out = Path(ctx.log_dir) / f"adjudicated_merged_{tag}.csv"
    sup_out = Path(ctx.log_dir) / f"adjudicated_superseded_{tag}.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(merged)
    with sup_out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields + ["superseded_by"],
                           extrasaction="ignore")
        w.writeheader()
        w.writerows(superseded)

    metrics: dict[str, float] = {
        "n_gold_in": float(len(gold_rows)), "n_round_in": float(len(round_rows)),
        "n_merged": float(len(merged)),
        "n_superseded_by_rules": float(len(superseded) - n_replaced),
        "n_replaced_by_round": float(n_replaced),
        "n_round_corrected": float(n_round_corrected),
    }
    for cid, n in sorted(per_check.items(), key=lambda kv: -kv[1]):
        metrics["superseded_" + re.sub(r"[^a-z0-9]+", "_", cid)] = float(n)
    msg = (f"{len(gold_rows)} gold + {len(round_rows)} {batch_label} -> "
           f"{len(merged)} merged; {len(superseded) - n_replaced} withheld as "
           f"rule-superseded, {n_replaced} replaced by newer review")
    return FunctionResult(ok=True,
                          outputs={"merged": str(out), "superseded": str(sup_out)},
                          metrics=metrics, message=msg)
