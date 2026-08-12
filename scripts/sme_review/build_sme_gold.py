#!/usr/bin/env python3
"""Build multi-reference SME gold assets from the batch-1 and batch-2 reviews.

Gold policy (agreed with the SME's own rules, Q&A 2026-08-09):
- Canonical gold = the SME correction when one exists (Partial/Wrong, and
  Acceptable rows where the SME still offered a preferred rendering);
  otherwise the reviewed translation (Correct/Acceptable).
- Alternates (``ko_all``, '|'-separated, canonical first): the reviewed
  translation stays an accepted alternate ONLY when the row was rated
  Correct/Acceptable (a preference, not an error); explicit synonyms the SME
  wrote in their notes ("synonym: ...") are added for batch-2.
- Batch-1 ``sme_independent_ko`` (the SME's from-scratch rendering) is an
  alternate when it differs from canonical.
- WRONG rows without a correction are dropped (no trustworthy gold).

Splits: stratified by (batch, rating) with a fixed seed. The TEST split is
held out from all optimisation forever; GEPA sees only train/dev.

Inputs (repo-relative):
  data/second_sme_review/second-100-term-subset-KO(with comments from HAPark).xlsx
  data/sme_review/2026-07-14/batch2_diverse_sent_100.csv   (what was sent)
  data/languages/ko/sme_review/2026-04-24/sme_labels_v1.csv (batch-1)

Outputs:
  data/evals/korean/sme_gold_splits/sme_gold_all.csv
  data/evals/korean/sme_gold_splits/{train,dev,test}.csv
  data/evals/korean/sme_gold_splits/audit.json
"""
from __future__ import annotations

import csv
import json
import random
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "data/evals/korean/sme_gold_splits"
XLSX = ROOT / "data/second_sme_review/second-100-term-subset-KO(with comments from HAPark).xlsx"
SENT = ROOT / "data/sme_review/2026-07-14/batch2_diverse_sent_100.csv"
BATCH1 = ROOT / "data/languages/ko/sme_review/2026-04-24/sme_labels_v1.csv"

SEED = 20260809
TEST_N, DEV_N = 40, 30  # remainder -> train

SYN_RE = re.compile(r"^synonyms?\s*[:=]\s*(.+)", re.IGNORECASE)
# The SME's DOMINANT way of saying "your rendering is acceptable, I just prefer
# mine": a free-text note (anywhere in the cell) reading "original translation
# can be used as a synonym". 18 of the 20 synonym-bearing batch-2 notes use
# this phrasing rather than a "synonym: X" line, so missing it scored SME-
# ACCEPTED renderings as failures and deflated every model's exact-match.
ORIGINAL_OK_RE = re.compile(
    r"orig(?:inal)?\s+translation.{0,40}?\bas\s+a\s+synonym", re.IGNORECASE | re.DOTALL)
PREFIX_RE = re.compile(r"^\s*preferred term:\s*", re.IGNORECASE)
OR_SPLIT_RE = re.compile(r"\s+or\s+|혹은")
# Consume a trailing 촬영 so "X선 촬영" -> "단순 촬영", not "단순 촬영 촬영".
XRAY_RE = re.compile(r"(일반\s*)?[xX]\s*-?\s*선(\s*촬영)?")


def _clean(s) -> str:
    return " ".join(str(s).split()) if s is not None else ""


def _token_overlap(a: str, b: str) -> float:
    ta, tb = set(a.split()), set(b.split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / min(len(ta), len(tb))


def parse_gold_refs(raw: str) -> tuple[list[str], list[str]]:
    """Split a free-text SME gold cell into clean reference strings.

    Batch-1 cells sometimes carry a literal "preferred term:" prefix, two
    alternatives joined by " or " / 혹은, a trailing comma, or two full
    alternates joined by ", " (only split on comma when the segments are
    long and share enough tokens to be renderings of the same term — commas
    also legitimately appear inside multi-site terms). Returns (refs, flags).
    """
    flags: list[str] = []
    g = _clean(raw)
    if PREFIX_RE.match(g):
        g = PREFIX_RE.sub("", g)
        flags.append("stripped_prefix")
    if g.endswith(","):
        g = g.rstrip(",").strip()
        flags.append("stripped_trailing_comma")
    parts = [p.strip() for p in OR_SPLIT_RE.split(g) if p.strip()]
    if len(parts) > 1:
        flags.append("split_or_alternates")
    refs: list[str] = []
    for part in parts:
        segs = [s.strip() for s in part.split(", ") if s.strip()]
        if (len(segs) == 2 and all(len(s.split()) >= 4 for s in segs)
                and _token_overlap(segs[0], segs[1]) >= 0.4):
            refs.extend(segs)
            flags.append("split_comma_alternates")
        else:
            refs.append(part)
    return [r for r in refs if r], flags


def _norm_rating(s: str) -> str:
    s = _clean(s).upper()
    return {"CORRECT": "CORRECT", "ACCEPTABLE": "ACCEPTABLE",
            "PARTIAL": "PARTIAL", "WRONG": "WRONG"}.get(s, "UNRATED")


def batch2_rows() -> list[dict]:
    sent = {r["sctid"].strip(): r for r in csv.DictReader(SENT.open(encoding="utf-8"))}
    ws = openpyxl.load_workbook(XLSX)["Sheet1"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sctid = _clean(row[idx["sctid"]])
        if not sctid:
            continue
        reviewed = _clean(row[idx["translation"]])
        rating = _norm_rating(row[idx["sme_rating"]])
        corrected = _clean(row[idx["sme_corrected_ko"]])
        notes = str(row[idx["sme_notes"]] or "")
        if corrected.lower() in ("", "ok", "nan"):
            corrected = ""
        if rating == "UNRATED" and corrected:
            rating = "PARTIAL"  # correction given but rating cell left blank
        if rating == "WRONG" and not corrected:
            continue
        gold = corrected or reviewed
        if not gold:
            continue
        alternates = []
        flags: list[str] = []
        if corrected and rating in ("CORRECT", "ACCEPTABLE") and reviewed != gold:
            alternates.append(reviewed)
        # SME explicitly blessed the reviewed rendering as an acceptable synonym.
        if ORIGINAL_OK_RE.search(notes) and reviewed and reviewed != gold \
                and reviewed not in alternates:
            alternates.append(reviewed)
            flags.append("sme_original_as_synonym")
        for line in notes.splitlines():
            m = SYN_RE.match(line.strip())
            if m:
                syn = _clean(m.group(1))
                if syn and syn not in alternates and syn != gold:
                    alternates.append(syn)
                    flags.append("sme_named_synonym")
        pt = _clean(row[idx["preferred_term"]]) or (
            sent[sctid]["preferred_term"] if sctid in sent else "")
        out.append({"sctid": sctid, "preferred_term": pt, "ko_reference": gold,
                    "ko_all": "|".join([gold] + alternates),
                    "reviewed_ko": reviewed,
                    "batch": "batch2", "sme_rating": rating,
                    "gold_flags": ";".join(sorted(set(flags)))})
    return out


def batch1_rows() -> list[dict]:
    out = []
    for r in csv.DictReader(BATCH1.open(encoding="utf-8")):
        sctid = _clean(r["sctid"])
        rating = _norm_rating(r.get("sme_rating", ""))
        reviewed = _clean(r.get("pipeline_translation_ko", ""))
        corrected_raw = _clean(r.get("sme_corrected_ko", ""))
        independent_raw = _clean(r.get("sme_independent_ko", ""))
        if corrected_raw.lower() in ("", "ok", "nan"):
            corrected_raw = ""
        if rating == "WRONG" and not (corrected_raw or independent_raw):
            continue
        flags: list[str] = []
        corrected, cf = parse_gold_refs(corrected_raw) if corrected_raw else ([], [])
        independent, if_ = parse_gold_refs(independent_raw) if independent_raw else ([], [])
        flags += cf + if_
        refs = corrected or independent or ([reviewed] if rating == "ACCEPTABLE" else [])
        if not refs:
            continue
        gold = refs[0]
        alternates = refs[1:]
        if rating == "ACCEPTABLE" and reviewed and reviewed != gold:
            alternates.append(reviewed)
        for ind in independent:
            if ind != gold and ind not in alternates:
                alternates.append(ind)
        # Batch-1 predates the batch-2 rulings. Add ruling-updated renderings as
        # accepted alternates so compliance with the SME's newer instructions is
        # not scored as a miss. Evidence the rulings are real: batch-2 gold
        # (written after them) contains ZERO 조영상 and no X선 forms.
        for ref in list([gold] + alternates):
            if XRAY_RE.search(ref):
                updated = _clean(XRAY_RE.sub("단순 촬영", ref))
                if updated not in ([gold] + alternates):
                    alternates.append(updated)
                if "pre_ruling_xray" not in flags:
                    flags.append("pre_ruling_xray")
            if "조영상" in ref:  # ruling: produced images end 영상, not 조영상
                updated = _clean(ref.replace("조영상", "영상"))
                if updated not in ([gold] + alternates):
                    alternates.append(updated)
                if "pre_ruling_joyeongsang" not in flags:
                    flags.append("pre_ruling_joyeongsang")
        out.append({"sctid": sctid, "preferred_term": _clean(r["english_term"]),
                    "ko_reference": gold, "ko_all": "|".join([gold] + alternates),
                    "reviewed_ko": reviewed,
                    "batch": "batch1", "sme_rating": rating,
                    "gold_flags": ";".join(sorted(set(flags)))})
    return out


# Forms the SME explicitly declared EQUALLY ACCEPTABLE in the 2026-08-09 Q&A.
# Encoding them as alternate references makes the metric measure what the SME
# actually accepts instead of which of two blessed variants they happened to
# type. Applied to every gold row, both batches.
EQUIVALENCES = [
    # "Procedures should end in '조영' or '조영술'" — terminal suffix only.
    (re.compile(r"조영술$"), "조영"),
    (re.compile(r"조영$"), "조영술"),
    # "When used as a test itself, it should be '투시 검사' or 투시술."
    (re.compile(r"투시\s*검사"), "투시술"),
    (re.compile(r"투시술"), "투시 검사"),
]


def expand_equivalences(refs: list[str]) -> list[str]:
    """Add SME-declared-equivalent surface variants of each reference."""
    out = list(refs)
    for ref in refs:
        for pattern, replacement in EQUIVALENCES:
            variant = _clean(pattern.sub(replacement, ref))
            if variant != ref and variant not in out:
                out.append(variant)
    return out


def main() -> None:
    rows = batch2_rows() + batch1_rows()
    # Encode the SME's explicit "either form is acceptable" rulings as extra
    # accepted references (see EQUIVALENCES).
    for r in rows:
        refs = [x for x in r["ko_all"].split("|") if x.strip()]
        expanded = expand_equivalences(refs)
        if len(expanded) > len(refs):
            r["ko_all"] = "|".join(expanded)
            r["gold_flags"] = ";".join(
                sorted(set(filter(None, r.get("gold_flags", "").split(";")))
                       | {"sme_equivalence_variants"}))
    # batch-2 supersedes batch-1 on sctid collisions (fresher review)
    seen: dict[str, dict] = {}
    for r in rows:
        seen.setdefault(r["sctid"], r)
    rows = list(seen.values())

    rng = random.Random(SEED)
    strata: dict[tuple, list[dict]] = defaultdict(list)
    for r in rows:
        strata[(r["batch"], r["sme_rating"])].append(r)
    for members in strata.values():
        rng.shuffle(members)

    test, dev, train = [], [], []
    # proportional draw per stratum so every (batch, rating) class is in each split
    total = len(rows)
    for key, members in sorted(strata.items()):
        n = len(members)
        n_test = round(TEST_N * n / total)
        n_dev = round(DEV_N * n / total)
        test += members[:n_test]
        dev += members[n_test:n_test + n_dev]
        train += members[n_test + n_dev:]

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cols = ["sctid", "preferred_term", "ko_reference", "ko_all", "reviewed_ko", "batch", "sme_rating", "gold_flags"]

    def dump(name: str, data: list[dict]) -> None:
        with (OUT_DIR / name).open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(sorted(data, key=lambda r: r["sctid"]))

    dump("sme_gold_all.csv", rows)
    dump("train.csv", train)
    dump("dev.csv", dev)
    dump("test.csv", test)

    audit = {
        "seed": SEED,
        "inputs": {"batch2_xlsx": str(XLSX.relative_to(ROOT)),
                   "batch2_sent": str(SENT.relative_to(ROOT)),
                   "batch1_labels": str(BATCH1.relative_to(ROOT))},
        "totals": {"all": len(rows), "train": len(train), "dev": len(dev),
                   "test": len(test)},
        "by_stratum": {f"{b}/{s}": len(m) for (b, s), m in sorted(strata.items())},
        "multi_ref_rows": sum(1 for r in rows if "|" in r["ko_all"]),
        "flagged_gold_rows": sum(1 for r in rows if r.get("gold_flags")),
        "test_sctids": sorted(r["sctid"] for r in test),
    }
    (OUT_DIR / "audit.json").write_text(
        json.dumps(audit, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({k: audit[k] for k in ("totals", "by_stratum", "multi_ref_rows")},
                     indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
